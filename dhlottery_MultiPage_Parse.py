####### Module Import #######
from selenium import webdriver
import openpyxl

driver=webdriver.Chrome('./chromedriver.exe')

################# Set openpyxl #####################
wb=openpyxl.load_workbook('./Test_1.xlsx', data_only=True)
sheet = wb['Sheet1']

#################### Set selenium & Chrome Driver API ##############
url_form = 'https://www.dhlottery.co.kr/gameResult.do?method=byWin&drwNo='
url_form_num = '855'

url = url_form + url_form_num

driver.get(url)

############### Main Code ####################
driver.implicitly_wait(time_to_wait=5)

data = []

round_element = driver.find_element_by_xpath('//*[@id="article"]/div[2]/div/div[2]/h4/strong')
xpath_element = driver.find_elements_by_xpath('//*[@id="article"]/div[2]/div/div[2]/div/div[1]/p/span')

for i in range(0, 6):
    data.append(xpath_element[i].text)
    # Win Number 1~6

data.append(driver.find_element_by_xpath('//*[@id="article"]/div[2]/div/div[2]/div/div[2]/p/span').text)
# + Bonus Number

print('---------- 회차 / 추첨번호 + 보너스번호 : ')

print(round_element.text + ' => ', end='')
print(data)

sheet.cell(row=2, column=1, value=round_element.text)

for i in range(2, 9):
    sheet.cell(row=2, column=i, value=int(data[i-2]))

wb.save('./Test_1.xlsx')


# Form Data 확인 후 쿼리에서 https://www.dhlottery.co.kr/gameResult.do?method=byWin&drwNo=944
# 하면 944 페이지로 이동됨
