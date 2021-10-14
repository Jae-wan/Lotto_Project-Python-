from selenium import webdriver
import openpyxl

def GET_DRIVER(driver, URL_SET):
    url = URL_SET
    global xl_row
    driver.get(url)
    data=[]

    round_element = driver.find_element_by_xpath('//*[@id="article"]/div[2]/div/div[2]/h4/strong')
    xpath_win_element = driver.find_elements_by_xpath('//*[@id="article"]/div[2]/div/div[2]/div/div[1]/p/span')
    xpath_bonus_element = driver.find_element_by_xpath('//*[@id="article"]/div[2]/div/div[2]/div/div[2]/p/span')

    for i in range(0, 6):
        data.append(xpath_win_element[i].text)
        # Win Number 1~6
    data.append(xpath_bonus_element.text)

    print(str(data))

    sheet.cell(row=xl_row, column=1, value=round_element.text) # 회차 기록용
    for c in range(2, 9):
        sheet.cell(row=xl_row, column=c, value=int(data[c-2])) # 번호 전체 기록용

    xl_row=xl_row + 1 # 행 바꾸기

# Scope : Global
xl_row = 2 # 엑셀 회차별로 한줄씩 내려갈려고

if __name__=="__main__":
    driver = webdriver.Chrome('./chromedriver.exe')
    URL_FRONT = 'https://www.dhlottery.co.kr/gameResult.do?method=byWin&drwNo='
    URL_BACK = 964

    wb=openpyxl.load_workbook('./Test_1.xlsx', data_only=True)
    sheet = wb['Sheet1']

    for i in range(URL_BACK, 930, -1) :
        URL_SET = URL_FRONT + str(i)

        print(str(i) + "에서 추출된 값 : ", end='')
        # GET_DRIVER의 str(data)를 이어서 출력. 단순히 결과 확인용

        GET_DRIVER(driver, URL_SET)

    wb.save('./Test_1.xlsx')

# Form Data 확인 후 쿼리에서 https://www.dhlottery.co.kr/gameResult.do?method=byWin&drwNo=944
# 하면 944 페이지로 이동됨
